from openpyxl import load_workbook

def leer_productos():
    wb = load_workbook("productos.xlsx")
    hoja = wb.active
    productos = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        producto = {
            "Id": fila[0],
            "Nombre": fila[1],
            "Precio": fila[2],
            "Cantidad": fila[3]
        }
        productos.append(producto)

    return productos